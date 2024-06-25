"""This module implements a Telegram bot for conducting a timed quiz.
The bot interacts with users by asking them questions,
collecting their responses, and evaluating their performance.

The bot performs the following main functions:
1. Loads quiz configuration and data from external files.
2. Handles user interactions via Telegram messages and commands.
3. Manages the quiz flow, including starting the test,
asking questions, and ending the test.
4. Saves the test results to an Excel file.

Configuration settings are loaded from a config.ini file,
and quiz data is read from separate files for
questions, options, answers, and candidate information.

Dependencies:
- openpyxl
- python-telegram-bot

To run this bot, execute the script and ensure the necessary
files and configurations are in place.
"""
import logging
import configparser
import datetime
import sys

import openpyxl
from openpyxl.styles import Font

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder, CommandHandler, ConversationHandler, MessageHandler,
    CallbackQueryHandler, filters, CallbackContext
)

# Load configuration from config.ini
config = configparser.ConfigParser()
config.read('config.ini')

# Load configuration
TIME_LIMIT = config.getint('time', 'time_limit')
QUESTIONS_PATH = config.get('file_paths', 'questions_file_path')
OPTIONS_PATH = config.get('file_paths', 'options_file_path')
ANSWERS_PATH = config.get('file_paths', 'answers_file_path')
CANDIDATE_INFO_PATH = config.get('file_paths', 'candidate_info_path')
EXCEL_FILE_PATH = config.get('file_paths', 'excel_file_path')
TEST_NAME = config.get('test_info', 'test_name')

# Bot token
TOKEN: str = config.get('BOT_ID', 'bot_token')

# Define conversation states
HANDLE_RESPONSE, START_TEST, ASK_QUESTION = range(3)


def format_time() -> str:
    """Format the time limit into a human-readable string.
    Returns:
        str: Formatted time string.
    """
    if TIME_LIMIT < 3600:
        minute = TIME_LIMIT / 60
        if minute < 1:
            return f"{int(TIME_LIMIT)} seconds"
        else:
            return f"{int(minute)} minutes"
    else:
        hours = TIME_LIMIT // 3600
        minutes = (TIME_LIMIT % 3600) // 60
        if minutes > 0:
            return f"{hours} hour and {minutes} minutes"
        else:
            return f"{hours} hours"


def load_files() -> None:
    """Read questions, options, and answers from text files"""
    global QUESTIONS, OPTIONS, ANSWERS, candidate_info
    try:
        with open(QUESTIONS_PATH, 'r', encoding="utf-8") as file:
            QUESTIONS = file.read().splitlines()

        with open(OPTIONS_PATH, 'r', encoding="utf-8") as file:
            OPTIONS = [line.split(',') for line in file.read().splitlines()]

        with open(ANSWERS_PATH, 'r', encoding="utf-8") as file:
            ANSWERS = file.read().splitlines()

        with open(CANDIDATE_INFO_PATH, 'r', encoding="utf-8") as file:
            candidate_info = {}
            for line in file.read().splitlines():
                exam_no, name = line.split(',')
                candidate_info[exam_no.strip()] = name.strip().upper()
    except Exception as e:
        print(f"one or more files may be missing {e}\nShutting down..")
        sys.exit(1)


def save_result(candidate_name,
                candidate_exam_no,
                total_questions,
                attempted_questions,
                unattempted_questions,
                failed_questions,
                final_score,
                grade) -> None:
    """Save the test result to an Excel file."""
    try:
        # Load the workbook and sheet
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        ws = wb.active

        # Check if the header needs to be written
        if ws['A1'].value is None:
            headers = ["Test Name", "Date", "Time", "Candidate Name", "Candidate Exam Number",
                       "Number of Questions", "Questions Attempted",
                       "Questions Unattempted",
                       "Questions Failed", "Final Score", "Grade"]
            ws.append(headers)

            # Set headers to bold and size 15
            for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
                for cell in col:
                    cell.font = Font(bold=True, size=15)

        # Add the test results
        data = [
            TEST_NAME,
            datetime.datetime.now().strftime("%Y-%m-%d"),
            datetime.datetime.now().strftime("%H:%M:%S"),
            candidate_name,
            candidate_exam_no,
            total_questions,
            attempted_questions,
            unattempted_questions,
            failed_questions,
            final_score,
            grade
        ]

        ws.append(data)
        wb.save(EXCEL_FILE_PATH)
    except Exception as e:
        print(f"Error saving test {e}")
        return None


async def start(update: Update, _) -> int:
    """Function to ask the user their exam number"""
    await update.message.reply_text("Enter your exam number")
    return HANDLE_RESPONSE


async def handle_response(update: Update, context: CallbackContext) -> int:
    """Handle the exam number input from the user."""
    candidate_exam_no = update.message.text.strip()
    candidate_name = candidate_info.get(candidate_exam_no)

    if not candidate_name:
        await update.message.reply_text(
            f"Exam number {candidate_exam_no} not found."
            f" Perhaps you are not enrolled for this test.")
        return ConversationHandler.END
    else:
        context.user_data['candidate_exam_no'] = candidate_exam_no
        context.user_data['candidate_name'] = candidate_name

        keyboard = [[InlineKeyboardButton("START", callback_data='START')]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(
            f"Hi {candidate_name}! Welcome to the {TEST_NAME} test."
            f"You have {len(QUESTIONS)} questions. "
            f"The duration of the test is {format_time()}.",
            reply_markup=reply_markup)
        return START_TEST


async def start_test(update: Update, context: CallbackContext) -> int:
    """Function to start the test"""
    query = update.callback_query
    await query.answer()
    if query.data == 'START':
        await query.edit_message_text("Your test has started. Good luck!")

        context.user_data['start_time'] = datetime.datetime.now()
        context.user_data['end_time'] = (
                datetime.datetime.now()
                + datetime.timedelta(seconds=TIME_LIMIT)
        )

        context.user_data['score'] = 0
        context.user_data['question_number'] = 0

        return await ask_question(update, context)


async def ask_question(update: Update, context: CallbackContext) -> int:
    """Function to ask a question"""
    question_number = context.user_data['question_number']
    end_time = context.user_data['end_time']

    current_time = datetime.datetime.now()
    remaining_time = (end_time - current_time).total_seconds()

    if remaining_time <= 0:
        await update.effective_message.reply_text("Time up! Your test has been submitted.")
        return await end_test(update, context)

    if remaining_time < (TIME_LIMIT / 100) * 50:
        if remaining_time < 60:
            await update.effective_message.reply_text(
                f"Hurry up you have {int(remaining_time)} seconds left\n"
            )
        elif remaining_time < 3600:
            await update.effective_message.reply_text(
                f"Hurry up you have {int(remaining_time / 60)}"
                f" minutes and {int(remaining_time % 60)} seconds left\n")

    if question_number < len(QUESTIONS):
        question = QUESTIONS[question_number]
        options = OPTIONS[question_number]

        keyboard = [
            [InlineKeyboardButton(options[0], callback_data=options[0].split(":")[0]),
             InlineKeyboardButton(options[1], callback_data=options[1].split(":")[0])],
            [InlineKeyboardButton(options[2], callback_data=options[2].split(":")[0]),
             InlineKeyboardButton(options[3], callback_data=options[3].split(":")[0])]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)

        await update.effective_message.reply_text(
            f"Question {question_number + 1}: {question}", reply_markup=reply_markup
        )
        return ASK_QUESTION
    else:
        return await end_test(update, context)


async def handle_answer(update: Update, context: CallbackContext) -> int:
    """Handle the answer to a question"""
    query = update.callback_query
    await query.answer()
    selected_option = query.data
    question_number = context.user_data['question_number']
    correct_answer = ANSWERS[question_number]

    if selected_option == correct_answer:
        context.user_data['score'] += 1
        await update.effective_message.reply_text("Correct!")
    if selected_option != correct_answer:
        await update.effective_message.reply_text(f"Wrong! The correct answer is: {correct_answer}")

    context.user_data['question_number'] += 1

    if context.user_data['question_number'] < len(QUESTIONS):
        return await ask_question(update, context)
    else:
        return await end_test(update, context)


async def end_test(update: Update, context: CallbackContext) -> int:
    """Function to end the test and show the result"""
    candidate_name = context.user_data['candidate_name']
    candidate_exam_no = context.user_data['candidate_exam_no']
    total_questions = len(QUESTIONS)
    attempted_questions = context.user_data['question_number']
    unattempted_questions = total_questions - attempted_questions
    score = context.user_data['score']
    failed_questions = attempted_questions - score

    if score >= 0.7 * total_questions:
        grade = "A"
    elif score >= 0.6 * total_questions:
        grade = "B"
    elif score >= 0.5 * total_questions:
        grade = "C"
    elif score >= 0.45 * total_questions:
        grade = "D"
    else:
        grade = "F"

    save_result(candidate_name, candidate_exam_no, total_questions, attempted_questions,
                unattempted_questions, failed_questions, score, grade)

    await update.effective_message.reply_text(
        f"Test Completed!\nYour score is {score}/{total_questions}.\nYour grade is {grade}.")

    return ConversationHandler.END


async def cancel(update: Update, _) -> int:
    """Function to cancel the conversation"""
    await update.message.reply_text("Test cancelled.")
    return ConversationHandler.END


def main():
    """Main function to start the bot"""
    # Load questions, options, and answers from files
    load_files()

    # Set up logging
    logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                        level=logging.INFO)

    # Create application
    app = ApplicationBuilder().token(TOKEN).build()

    # Create conversation handler
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler('start', start)],
        states={
            HANDLE_RESPONSE: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_response)],
            START_TEST: [CallbackQueryHandler(start_test)],
            ASK_QUESTION: [CallbackQueryHandler(handle_answer)]
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )

    # Add handlers
    app.add_handler(conv_handler)

    # Print message indicating the bot has started.
    print("Bot Started...")

    # Start the bot
    app.run_polling()


if __name__ == "__main__":
    main()
