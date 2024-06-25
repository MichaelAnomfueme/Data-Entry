import configparser
import time
import datetime
import openpyxl
from openpyxl.styles import Font

# Read configuration from config.ini
config = configparser.ConfigParser()
config.read('config.ini')

# load configuration
TIME_LIMIT = config.getint('time', 'time_limit')
QUESTIONS_PATH = config.get('file_paths', 'questions_file_path')
OPTIONS_PATH = config.get('file_paths', 'options_file_path')
ANSWERS_PATH = config.get('file_paths', 'answers_file_path')
CANDIDATE_INFO_PATH = config.get('file_paths', 'candidate_info_path')
EXCEL_FILE_PATH = config.get('file_paths', 'excel_file_path')
TEST_NAME = config.get('test_info', 'test_name')


def format_time() -> str:
    if TIME_LIMIT < 3600:
        minute = TIME_LIMIT / 60
        if minute < 1:
            return f"{int(TIME_LIMIT)} seconds"
        else:
            return f"{int(minute)} minutes"
    else:
        hours = TIME_LIMIT // 3600
        minutes = (TIME_LIMIT % 3600) // 60
        if minutes > 0:
            return f"{hours} hour and {minutes} minutes"
        else:
            return f"{hours} hours"


# Read questions, options, and answers from text files
def load_files() -> None:
    global QUESTIONS, OPTIONS, ANSWERS, EXAM_NO, candidate_info
    try:
        with open(QUESTIONS_PATH, 'r') as file:
            QUESTIONS = file.read().splitlines()

        with open(OPTIONS_PATH, 'r') as file:
            OPTIONS = [line.split(',') for line in file.read().splitlines()]

        with open(ANSWERS_PATH, 'r') as file:
            ANSWERS = file.read().splitlines()

        with open(ANSWERS_PATH, 'r') as file:
            ANSWERS = file.read().splitlines()

        with open(CANDIDATE_INFO_PATH, 'r') as file:
            candidate_info = {}
            for line in file.read().splitlines():
                exam_no, name = line.split(',')
                candidate_info[exam_no.strip()] = name.strip().upper()
    except Exception as e:
        print(f"one or more files may be missing {e}\nShutting down..")
        exit(1)


def save_result(candidate_name,
                candidate_exam_no,
                total_questions,
                attempted_questions,
                unattempted_questions,
                failed_questions,
                final_score, grade) -> None:
    try:
        # Load the workbook and sheet
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        ws = wb.active

        # Check if the header needs to be written
        if ws['A1'].value is None:
            headers = ["Test Name", "Date", "Time", "Candidate Name", "Candidate Exam Number",
                       "Number of Questions", "Questions Attempted",
                       "Questions Unattempted",
                       "Questions Failed", "Final Score", "Grade"]
            ws.append(headers)

            # Set headers to bold and size 15
            for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
                for cell in col:
                    cell.font = Font(bold=True, size=15)

        # Add the test results
        data = [
            TEST_NAME,
            datetime.datetime.now().strftime("%Y-%m-%d"),
            datetime.datetime.now().strftime("%H:%M:%S"),
            candidate_name,
            candidate_exam_no,
            total_questions,
            attempted_questions,
            unattempted_questions,
            failed_questions,
            final_score,
            grade
        ]

        ws.append(data)
        wb.save(EXCEL_FILE_PATH)
    except Exception as e:
        print(f"Error saving test {e}")
        return None


def main() -> None:
    """Main function to run the quiz."""
    try:
        global remaining_time
        format_time()
        load_files()
        candidate_exam_no: str = input("Enter your exam number: ").strip()
        candidate_name = candidate_info.get(candidate_exam_no)

        if not candidate_name:
            print(f"Exam number {candidate_exam_no} not found. Perhaps you are not enrolled for this test.\n")
            return None
        else:
            print(f"\nHi {candidate_name}! Welcome to the {TEST_NAME} test. "
                  f"You have {len(QUESTIONS)} questions with 2 attempts each. "
                  f"The duration of the test is {format_time()}\n")

        start_test = input("If you wish to begin the test enter 'START' ").upper()

        if start_test == "START":
            print("Your test has started. Good luck!\n")
        else:
            print("You have made an invalid input restart the test and enter 'START' to begin test. ")
            return None

        start_time = time.time()  # Record the start time
        end_time = start_time + TIME_LIMIT  # Calculate the end time
        score: int = 0
        question_number: int = 1

        for i in range(len(QUESTIONS)):
            current_time = time.time()
            remaining_time = end_time - current_time
            if remaining_time <= 0:
                print("Time up! Your test has been automatically submitted.")
                break

            if remaining_time < (TIME_LIMIT / 100) * 30:
                if remaining_time < 60:
                    print(f"Hurry up you have {int(remaining_time)} seconds left\n")
                elif remaining_time < 3600:
                    print(f"Hurry up you have {int(remaining_time / 60)} minutes and "
                          f"{int(remaining_time % 60)} seconds left\n")

            print(QUESTIONS[i])
            for option in OPTIONS[i]:
                print(option)

            for attempt in range(2):
                user_answer = input("\nAnswer: ").upper()
                if user_answer == ANSWERS[i]:
                    print("Congratulations! You are correct!")
                    score += 1
                    print(f"Questions attempted = {question_number}/{len(QUESTIONS)}")
                    print(f"Score = {score}/{len(QUESTIONS)}\n")
                    break
                else:
                    if attempt == 0:
                        print("Incorrect! You have one attempt left.")
                    else:
                        print(f"Oops! You failed question {question_number}. The correct answer is {ANSWERS[i]}")
                        print(f"Questions attempted = {question_number}/{len(QUESTIONS)}")
                        print(f"Score = {score}/{len(QUESTIONS)}\n")

            question_number += 1

        if remaining_time > 0:
            print("You have completed your test.")

        print(f"Your score is {score}/{len(QUESTIONS)}")

        if score >= 0.7 * len(QUESTIONS):
            grade = "A"
        elif score >= 0.6 * len(QUESTIONS):
            grade = "B"
        elif score >= 0.5 * len(QUESTIONS):
            grade = "C"
        elif score >= 0.45 * len(QUESTIONS):
            grade = "D"
        else:
            grade = "F"

        print(f"Your grade is: {grade}")

        # Calculate unattempted questions
        unattempted_questions = len(QUESTIONS) - question_number + 1
        failed_questions = question_number - 1 - score

        # Save the results to an Excel file
        save_result(candidate_name,
                    candidate_exam_no,
                    len(QUESTIONS),
                    question_number - 1,
                    unattempted_questions,
                    failed_questions,
                    score,
                    grade)

    except Exception as e:
        print(f"Error handing test {e}")
        return None


if __name__ == "__main__":
    main()
